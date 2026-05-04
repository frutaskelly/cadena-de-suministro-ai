"""Procesador del Excel BD del cliente EHMO (port v1 simplificado).

Toma un .xlsx con hoja "BD" (formato EHMO) y:
1. Lee filas con UNIDAD, LOTE, ALIMENTO, PRESENTACION, CANTIDAD.
2. Filtra excluidos por geografia + por palabra clave (lote 5 / FyV).
3. Aplica reglas de cambio (lote 1 -> lote 5) y exclusion fina (`is_ignorar`).
4. Construye filas de pedido y las pasa a `services.pedidos.from_batch_rows()`
   que las matchea contra catalogo + crea Pedidos en DB.
5. Devuelve resultado con pedidos creados + ruta a Excel original.

Es una version condensada del v1 (~840 LOC -> ~250 LOC) que:
- Reusa `services.pedidos.from_batch_rows()` para creacion + fuzzy match.
- Reusa `services.pdf_generators` para PDFs.
- Sube todo a Drive con `services.drive.upload_file()`.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Optional
from uuid import UUID

import openpyxl
from sqlalchemy.orm import Session

from .pedidos import PedidoRowIn, from_batch_rows
from .pdf_generators import (
    generar_lista_compras_pdf, generar_lista_compras_xlsx, generar_pedido_pdf,
)
from .drive import upload_file

log = logging.getLogger(__name__)

# Palabras clave que marcan unidades excluidas (no se les surte FyV)
EXCLUIDOS_KW = [
    "pichucalco", "palenque", "tila", "reforma",
    "yajalón", "yajalon", "amatán", "amatan",
]

# Productos que NO van en FyV aunque vengan marcados (mal clasificacion del cliente)
IGNORAR_KW = [
    "salchicha",
    "queso oaxaca", "queso panela", "queso fresco", "queso amarillo",
    "yogurt", "yoghurt",
    "huevo", "huevos",
]

# Productos que vienen en Lote 1 pero deben moverse a Lote 5 (cambio confirmado)
CAMBIO_KW = [
    "ajo en bulbo", "ajonjolí", "ajonjoli",
    "cacahuate tostado sin sal",
    "canela en raja",
    "chile seco ancho", "chile seco guajillo", "chile seco pasilla",
    "epazote", "flor de jamaica",
    "nuez sin cascara", "nuez sin cáscara",
    "orégano en hoja", "oregano en hoja",
    "perejil",
    "te de limón zacate", "te de limon zacate",
    "te de manzanilla", "té de manzanilla",
    "te de yerbabuena", "té de yerbabuena",
]

MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}


@dataclass
class ExcelBDResult:
    fecha_iso: str
    fecha_legible: str
    pedidos_creados: list[dict] = field(default_factory=list)
    pedidos_skipped: list[dict] = field(default_factory=list)
    unidades_sin_match: list[str] = field(default_factory=list)
    lineas_sin_match: list[dict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    # archivos generados (rutas locales y URLs Drive)
    pedido_pdf_path: Optional[str] = None
    pedido_pdf_drive_url: Optional[str] = None
    lista_compras_pdf_path: Optional[str] = None
    lista_compras_pdf_drive_url: Optional[str] = None
    lista_compras_xlsx_path: Optional[str] = None
    lista_compras_xlsx_drive_url: Optional[str] = None


def _is_excluido(unidad: str) -> bool:
    n = (unidad or "").lower()
    return any(kw in n for kw in EXCLUIDOS_KW)


def _is_ignorar(alimento: str) -> bool:
    a = (alimento or "").lower()
    return any(kw in a for kw in IGNORAR_KW)


def _is_cambio(alimento: str) -> bool:
    a = (alimento or "").lower()
    if _is_ignorar(a):
        return False
    return any(kw in a for kw in CAMBIO_KW)


def _es_lote_5(lote_str) -> bool:
    s = str(lote_str or "").strip().upper()
    return s in ("5 FRUTAS Y VERDURAS", "FRUTAS Y VERDURAS")


def _es_lote_1(lote_str) -> bool:
    s = str(lote_str or "").strip().upper()
    return s in ("1 ABARROTES", "ABARROTES")


def _extraer_fecha(filename: str, excel_path: Optional[Path] = None) -> tuple[date, str]:
    """Devuelve (fecha_date, "27 de abril")."""
    today = date.today()

    # 1) Buscar en celda A3 del Excel (formato "Suma de 28-abr")
    if excel_path:
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for r in range(1, 6):
                    for c in range(1, 5):
                        val = ws.cell(row=r, column=c).value
                        if not val or not isinstance(val, str):
                            continue
                        m = re.search(r"(\d{1,2})[\s\-/](\w{3,})", val.lower())
                        if m and "suma" in val.lower():
                            dia = int(m.group(1))
                            mes_prefijo = m.group(2)[:3]
                            for nombre, num in MESES.items():
                                if nombre.startswith(mes_prefijo):
                                    wb.close()
                                    return (
                                        date(today.year, num, dia),
                                        f"{dia} de {nombre}",
                                    )
            wb.close()
        except Exception:
            pass

    # 2) Buscar fecha YYYY-MM-DD en filename
    m = re.search(r"(20\d{2})-(\d{2})-(\d{2})", filename or "")
    if m:
        try:
            d = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            mes = list(MESES.keys())[d.month - 1]
            return (d, f"{d.day} de {mes}")
        except ValueError:
            pass

    # 3) "27 de abril" o "27_de_abril" en filename
    fname_norm = (filename or "").replace("_", " ").lower()
    m = re.search(r"(\d{1,2})\s+de\s+(\w+)", fname_norm)
    if m:
        dia = int(m.group(1))
        mes_kw = m.group(2)[:3]
        for nombre, num in MESES.items():
            if nombre.startswith(mes_kw):
                return (date(today.year, num, dia), f"{dia} de {nombre}")

    # 4) Hoy
    mes = list(MESES.keys())[today.month - 1]
    return (today, f"{today.day} de {mes}")


def procesar_excel_bd(
    db: Session,
    tenant_id: UUID,
    excel_path: Path,
    *,
    canal: str = "EXCEL_BD",
    cliente_id: Optional[UUID] = None,
    contrato_id: Optional[UUID] = None,
    output_dir: Optional[Path] = None,
    upload_drive: bool = True,
) -> ExcelBDResult:
    """Procesa un Excel BD y crea pedidos + genera PDFs.

    1. Lee la hoja BD del Excel.
    2. Filtra (lote 5, no excluidos, no ignorados; aplica cambios lote1->lote5).
    3. Llama from_batch_rows() para crear Pedidos en DB.
    4. Genera PDFs (pedido por hospital + lista compras + xlsx).
    5. Sube a Drive (si configurado).
    """
    excel_path = Path(excel_path)
    fecha, fecha_legible = _extraer_fecha(excel_path.name, excel_path)
    fecha_iso = fecha.isoformat()
    output_dir = Path(output_dir or f"/tmp/cadena_excel_bd/{fecha_iso}")
    output_dir.mkdir(parents=True, exist_ok=True)

    result = ExcelBDResult(fecha_iso=fecha_iso, fecha_legible=fecha_legible)

    # ─── 1) Leer Excel BD ─────────────────────────────────────────────
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if "BD" not in wb.sheetnames:
        result.warnings.append(
            f"El Excel no tiene hoja 'BD'. Hojas: {wb.sheetnames}"
        )
        return result

    ws = wb["BD"]
    headers = [str(c.value or "").upper().strip() for c in ws[1]]

    def col(name: str) -> Optional[int]:
        for i, h in enumerate(headers):
            if name in h:
                return i
        return None

    col_unidad = col("UNIDAD")
    col_lote = col("LOTE")
    col_alimento = col("ALIMENTO")
    col_presentacion = col("PRESENTACION") or col("PRESENTACIÓN")
    col_fecha = col("FECHA")
    # cantidad: el archivo de EHMO tiene varias columnas posibles,
    # tomamos la primera que parezca cantidad
    col_cantidad = col("CANTIDAD") or col("SUMA")
    col_cba = col("C.B.A")

    if col_unidad is None or col_alimento is None:
        result.warnings.append(
            f"Hoja BD sin columnas requeridas (UNIDAD/ALIMENTO). Headers: {headers}"
        )
        return result

    rows: list[PedidoRowIn] = []
    excluidos = 0
    ignorados = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[col_unidad]:
            continue
        unidad = str(row[col_unidad] or "").strip()
        lote = row[col_lote] if col_lote is not None else ""
        alimento = str(row[col_alimento] or "").strip()
        presentacion = (
            str(row[col_presentacion] or "").strip()
            if col_presentacion is not None
            else "KILO"
        ) or "KILO"
        cantidad = row[col_cantidad] if col_cantidad is not None else None
        cba = (
            str(row[col_cba] or "").strip()
            if col_cba is not None
            else None
        )

        if not alimento or cantidad is None:
            continue

        if _is_excluido(unidad):
            excluidos += 1
            continue

        if _is_ignorar(alimento):
            ignorados += 1
            continue

        # Aceptar lote 5 directo o cambio desde lote 1
        if not (_es_lote_5(lote) or (_es_lote_1(lote) and _is_cambio(alimento))):
            continue

        try:
            qty = Decimal(str(cantidad))
            if qty <= 0:
                continue
        except Exception:
            continue

        rows.append(PedidoRowIn(
            unidad_nombre=unidad,
            alimento=alimento,
            cantidad=qty,
            presentacion=presentacion.upper(),
            lote=str(lote) if lote else None,
            cba=cba or None,
        ))

    wb.close()
    log.info(
        f"Excel BD: {len(rows)} filas validas (excluidos={excluidos}, "
        f"ignorados={ignorados})"
    )
    if not rows:
        result.warnings.append(
            "No se encontraron filas validas (lote 5 + unidad valida)."
        )
        return result

    # ─── 2) Crear pedidos en DB usando servicio existente ────────────
    batch = from_batch_rows(
        db=db,
        tenant_id=tenant_id,
        fecha=fecha,
        rows=rows,
        canal=canal,
        contrato_id=contrato_id,
        cliente_id=cliente_id,
        force_overwrite=False,
        raw_payload={
            "source": "excel_bd",
            "filename": excel_path.name,
            "fecha_legible": fecha_legible,
        },
    )

    result.pedidos_creados = [
        {
            "pedido_id": str(p.pedido_id),
            "folio_interno": p.folio_interno,
            "unidad_nombre": p.unidad_nombre,
            "lineas_count": p.lineas_count,
            "total": float(p.total),
            "requires_review": p.requires_review,
        }
        for p in batch.pedidos_creados
    ]
    result.pedidos_skipped = batch.pedidos_skipped
    result.unidades_sin_match = batch.unidades_sin_match
    result.lineas_sin_match = batch.lineas_sin_match
    result.warnings.extend(batch.warnings)

    # Siempre generamos PDFs si tenemos rows validos. Aun si los pedidos
    # ya existian (skipped), regenerar los PDFs es util para el operador.
    if not rows:
        return result

    # ─── 3) Generar PDFs/XLSX desde rows agrupadas ──────────────────
    items_pdf = [
        {
            "unidad": r.unidad_nombre,
            "alimento": r.alimento,
            "presentacion": r.presentacion,
            "cantidad": float(r.cantidad),
        }
        for r in rows
    ]

    pedido_pdf = output_dir / f"Pedido {fecha_iso} Frutas y Verduras.pdf"
    lista_pdf = output_dir / f"Lista de Compras {fecha_iso}.pdf"
    lista_xlsx = output_dir / f"Lista de Compras {fecha_iso}.xlsx"

    try:
        generar_pedido_pdf(items_pdf, fecha_legible, pedido_pdf)
        result.pedido_pdf_path = str(pedido_pdf)
    except Exception as e:
        result.warnings.append(f"PDF pedido fallo: {e}")

    try:
        generar_lista_compras_pdf(items_pdf, fecha_legible, lista_pdf)
        result.lista_compras_pdf_path = str(lista_pdf)
    except Exception as e:
        result.warnings.append(f"Lista compras PDF fallo: {e}")

    try:
        generar_lista_compras_xlsx(items_pdf, fecha_legible, lista_xlsx)
        result.lista_compras_xlsx_path = str(lista_xlsx)
    except Exception as e:
        result.warnings.append(f"Lista compras XLSX fallo: {e}")

    # ─── 4) Subir a Drive ────────────────────────────────────────────
    if upload_drive:
        subfolder = fecha_iso  # subcarpeta por fecha
        for path_attr, url_attr in [
            ("pedido_pdf_path", "pedido_pdf_drive_url"),
            ("lista_compras_pdf_path", "lista_compras_pdf_drive_url"),
            ("lista_compras_xlsx_path", "lista_compras_xlsx_drive_url"),
        ]:
            local = getattr(result, path_attr)
            if not local:
                continue
            try:
                drive_resp = upload_file(Path(local), subfolder=subfolder)
                if drive_resp:
                    setattr(result, url_attr, drive_resp.get("link"))
            except Exception as e:
                result.warnings.append(f"Drive upload {path_attr}: {e}")

    return result
