"""Relacion consolidada de notas de remision del dia (port v1).

Genera 1 PDF + 1 Excel con todas las remisiones FACTURADA/CONFIRMADA
del dia agrupadas por cliente. Util para administracion + bancos.
"""
from __future__ import annotations

import logging
from datetime import date as date_cls
from decimal import Decimal
from pathlib import Path
from typing import Iterable
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)
from sqlalchemy.orm import Session

from ..models import Cliente, Remision, UnidadEntrega

log = logging.getLogger(__name__)

NEGRO = colors.HexColor("#000000")
GRIS_CLARO = colors.HexColor("#F4F4F4")
AZUL_OSC = colors.HexColor("#1F4E79")


def _fmt_money(v) -> str:
    try:
        return f"${float(v):,.2f}"
    except (TypeError, ValueError):
        return str(v)


def _query_remisiones(
    db: Session,
    tenant_id: UUID,
    fecha: date_cls,
    estados: Iterable[str] = ("CONFIRMADA", "FACTURADA"),
) -> list[tuple[Remision, Cliente, UnidadEntrega | None]]:
    rows = (
        db.query(Remision, Cliente, UnidadEntrega)
        .join(Cliente, Cliente.id == Remision.cliente_id)
        .outerjoin(UnidadEntrega, UnidadEntrega.id == Remision.unidad_entrega_id)
        .filter(
            Remision.tenant_id == tenant_id,
            Remision.fecha_generada == fecha,
            Remision.estado.in_(list(estados)),
        )
        .order_by(Cliente.codigo, Remision.folio)
        .all()
    )
    return rows


def generar_relacion_pdf(
    db: Session,
    tenant_id: UUID,
    fecha: date_cls,
    output_path: Path,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = _query_remisiones(db, tenant_id, fecha)

    doc = SimpleDocTemplate(
        str(output_path), pagesize=letter,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        title=f"Relacion {fecha.isoformat()}",
    )

    styles = getSampleStyleSheet()
    style_h1 = ParagraphStyle(
        "h1", parent=styles["Heading1"],
        fontSize=14, leading=18,
        textColor=AZUL_OSC, alignment=TA_CENTER,
        spaceAfter=10,
    )
    style_sub = ParagraphStyle(
        "sub", parent=styles["Normal"],
        fontSize=10, leading=12,
        textColor=colors.HexColor("#555555"),
        alignment=TA_CENTER, spaceAfter=14,
    )
    style_grupo = ParagraphStyle(
        "g", parent=styles["Normal"],
        fontSize=11, leading=13,
        fontName="Helvetica-Bold", spaceBefore=10, spaceAfter=4,
    )

    elements = [
        Paragraph(f"Relación de Documentos — {fecha.isoformat()}", style_h1),
        Paragraph(
            f"{len(rows)} notas de remisión emitidas",
            style_sub,
        ),
    ]

    if not rows:
        elements.append(Paragraph(
            "Sin remisiones para esta fecha.", styles["Normal"]
        ))
        doc.build(elements)
        return output_path

    by_cliente: dict[str, list] = {}
    for rem, cli, unidad in rows:
        key = f"{cli.codigo} — {cli.legal_name}"
        by_cliente.setdefault(key, []).append((rem, cli, unidad))

    total_general = Decimal(0)
    for cli_label, group in by_cliente.items():
        elements.append(Paragraph(cli_label, style_grupo))

        data = [["Folio", "Destino", "Estado", "Líneas", "Total"]]
        subtotal = Decimal(0)
        for rem, cli, unidad in group:
            data.append([
                rem.folio,
                (unidad.nombre if unidad else "—")[:50],
                rem.estado,
                str(len(rem.lineas) if rem.lineas else 0),
                _fmt_money(rem.total or 0),
            ])
            subtotal += rem.total or Decimal(0)
        data.append(["", "", "", "Subtotal:", _fmt_money(subtotal)])
        total_general += subtotal

        tbl = Table(
            data,
            colWidths=[2.5 * cm, 8 * cm, 2.5 * cm, 2 * cm, 2.5 * cm],
            repeatRows=1,
        )
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), AZUL_OSC),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (3, 0), (4, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -2), 0.3, colors.HexColor("#BBBBBB")),
            ("BACKGROUND", (0, -1), (-1, -1), GRIS_CLARO),
            ("FONTNAME", (3, -1), (-1, -1), "Helvetica-Bold"),
        ]))
        elements.append(tbl)

    elements.append(Spacer(1, 0.3 * cm))
    style_tg = ParagraphStyle(
        "tg", parent=styles["Normal"],
        fontSize=12, leading=14,
        fontName="Helvetica-Bold",
        alignment=TA_RIGHT,
    )
    elements.append(
        Paragraph(
            f"Total general del día: <b>{_fmt_money(total_general)}</b>",
            style_tg,
        )
    )
    doc.build(elements)
    log.info(f"Relacion PDF: {output_path.name} ({len(rows)} remisiones)")
    return output_path


def generar_relacion_xlsx(
    db: Session,
    tenant_id: UUID,
    fecha: date_cls,
    output_path: Path,
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = _query_remisiones(db, tenant_id, fecha)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Relacion {fecha.isoformat()}"

    bold_white = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    border = Border(
        left=Side(border_style="thin", color="BBBBBB"),
        right=Side(border_style="thin", color="BBBBBB"),
        top=Side(border_style="thin", color="BBBBBB"),
        bottom=Side(border_style="thin", color="BBBBBB"),
    )

    ws["A1"] = f"Relación de Documentos — {fecha.isoformat()}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 28

    headers = ["Cliente", "Folio", "Destino", "Estado", "Líneas", "Total"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = bold_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    total_general = Decimal(0)
    for i, (rem, cli, unidad) in enumerate(rows, 1):
        r = 3 + i
        ws.cell(row=r, column=1, value=cli.legal_name).border = border
        ws.cell(row=r, column=2, value=rem.folio).border = border
        ws.cell(row=r, column=3, value=unidad.nombre if unidad else "").border = border
        ws.cell(row=r, column=4, value=rem.estado).border = border
        ws.cell(row=r, column=5, value=len(rem.lineas) if rem.lineas else 0).border = border
        c = ws.cell(row=r, column=6, value=float(rem.total or 0))
        c.number_format = '"$"#,##0.00'
        c.border = border
        total_general += rem.total or Decimal(0)

    last = 3 + len(rows) + 1
    ws.cell(row=last, column=1, value="TOTAL DEL DÍA").font = Font(bold=True)
    c_total = ws.cell(row=last, column=6, value=float(total_general))
    c_total.font = Font(bold=True)
    c_total.number_format = '"$"#,##0.00'
    for col in range(1, 7):
        ws.cell(row=last, column=col).fill = PatternFill(
            start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"
        )
        ws.cell(row=last, column=col).border = border

    widths = {1: 35, 2: 14, 3: 35, 4: 12, 5: 8, 6: 14}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    wb.save(str(output_path))
    log.info(f"Relacion XLSX: {output_path.name}")
    return output_path
