"""Builder de Base Maestra (Sprint 13).

Toma un folder con subcarpetas de hospitales (cada una con 1+ archivos
.xlsx) y produce el archivo consolidado Base Maestra con 5 hojas:
- BD: data raw consolidada (UNIDAD x SKU x dia)
- Almacen: pivot LOTE 1 ABARROTES (suma semanal)
- Carne: pivot LOTE 2 PROTEINAS
- FyV: pivot LOTE 5 FRUTAS Y VERDURAS
- Hoja1: pivot completo

Bullet-proof:
- Procesa cada hospital independientemente; falla aislada no aborta el batch
- Si un hospital tiene varios archivos, los mergea
- Reporta hospitales OK / WARN / FAIL al final
- Comparacion contra base maestra anterior si se proporciona
"""
from __future__ import annotations

import logging
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .base_maestra_parser import parse_pedido_hospital

log = logging.getLogger(__name__)

DAYS_OF_WEEK_ES = ["LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM"]


@dataclass
class HospitalParseResult:
    folder: Path
    nombre: str
    nombre_canonico: Optional[str]
    estado: str  # OK | WARN | FAIL
    archivos: list[str]
    items: list[dict]
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    hojas_usadas: list[str] = field(default_factory=list)


@dataclass
class BaseMaestraResult:
    fecha_inicio: date
    fecha_fin: date
    semana_label: str
    hospitales: list[HospitalParseResult]
    output_path: Optional[Path] = None
    output_size: int = 0
    filas_bd: int = 0
    diff_pct_vs_anterior: Optional[float] = None
    diff_resumen: Optional[dict] = None


def _norm(s: str) -> str:
    s = (s or "").strip().upper()
    return "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )


def _merge_items(items_a: list[dict], items_b: list[dict]) -> list[dict]:
    """Fusiona dos listas de items por (lote, alimento, presentacion).

    Si hay duplicados, suma las cantidades por dia. Esto permite que un
    hospital con varios archivos (regular + DESECHABLES + EXTRAS) quede
    en una sola lista consolidada.
    """
    by_key: dict[tuple, dict] = {}
    for source in (items_a, items_b):
        for it in source:
            key = (
                _norm(it.get("lote", "")),
                _norm(it.get("alimento", "")),
                _norm(it.get("presentacion", "")),
            )
            existing = by_key.get(key)
            if not existing:
                by_key[key] = dict(it)
                # Aseguramos copia del dict de dias
                by_key[key]["dias"] = dict(it.get("dias") or {})
                continue
            for d, q in (it.get("dias") or {}).items():
                existing["dias"][d] = (existing["dias"].get(d, 0) or 0) + q
            existing["total"] = sum(existing["dias"].values())
            # mantener precio del primero a menos que el segundo lo tenga
            if not existing.get("precio") and it.get("precio"):
                existing["precio"] = it["precio"]
    return list(by_key.values())


def _process_hospital_folder(
    folder: Path,
    fecha_inicio: date,
    fecha_fin: date,
) -> HospitalParseResult:
    archivos = sorted([
        p for p in folder.glob("*.xlsx") if not p.name.startswith("~$")
    ])
    res = HospitalParseResult(
        folder=folder,
        nombre=folder.name,
        nombre_canonico=None,
        estado="FAIL",
        archivos=[a.name for a in archivos],
        items=[],
    )
    if not archivos:
        # Tambien puede haber pdfs (H. Acala) que aun no soportamos
        pdfs = list(folder.glob("*.pdf"))
        if pdfs:
            res.errors.append(
                f"Solo PDFs encontrados ({len(pdfs)}). Soporte OCR pendiente."
            )
        else:
            res.errors.append("Carpeta vacia, sin xlsx")
        return res

    all_items: list[dict] = []
    hospital_canonico = None
    estados_archivos = []
    for archivo in archivos:
        try:
            r = parse_pedido_hospital(
                archivo, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin
            )
        except Exception as e:
            res.errors.append(f"[{archivo.name}] excepcion: {e}")
            continue

        estados_archivos.append(r["estado"])
        res.hojas_usadas.append(f"{archivo.name}:{r.get('hoja_usada')}")
        for w in r.get("warnings", []):
            res.warnings.append(f"[{archivo.name}] {w}")
        for e in r.get("errors", []):
            res.errors.append(f"[{archivo.name}] {e}")

        if r.get("hospital_nombre_archivo") and not hospital_canonico:
            hospital_canonico = r["hospital_nombre_archivo"]

        all_items = _merge_items(all_items, r.get("items") or [])

    res.nombre_canonico = hospital_canonico or folder.name
    res.items = all_items

    # Definir estado del hospital
    if not all_items:
        res.estado = "FAIL"
        res.errors.append("Sin items con cantidad > 0 en ningun archivo")
    elif res.warnings or "WARN" in estados_archivos:
        res.estado = "WARN"
    else:
        res.estado = "OK"
    return res


def _make_bd_rows(
    hospitales: list[HospitalParseResult],
    fecha_inicio: date,
    fecha_fin: date,
) -> list[dict]:
    """Genera las filas finales de la hoja BD."""
    rows = []
    delta = (fecha_fin - fecha_inicio).days + 1
    days = [fecha_inicio + timedelta(days=i) for i in range(delta)]

    for h in hospitales:
        if not h.items:
            continue
        for it in h.items:
            row = {
                "UNIDAD": h.nombre_canonico or h.nombre,
                "LOTE": it.get("lote", ""),
                "C.B.A": it.get("cba", ""),
                "ALIMENTO": it.get("alimento", ""),
                "PRESENTACIÓN": it.get("presentacion", ""),
                "PRECIO UNITARIO": it.get("precio") or 0,
            }
            total = 0.0
            for d in days:
                qty = (it.get("dias") or {}).get(d.isoformat())
                row[d.isoformat()] = qty if qty is not None else ""
                if qty:
                    total += qty
            row["TOTAL"] = total if total > 0 else (it.get("total") or "")
            row["Observaciones"] = ""
            rows.append(row)
    return rows


# ─── Excel Writer ──────────────────────────────────────────────────────
def _bold():
    return Font(name="Calibri", size=11, bold=True, color="FFFFFF")


def _header_fill():
    return PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")


def _border():
    s = Side(border_style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_bd_sheet(ws, rows: list[dict], days: list[date]):
    headers = (
        ["UNIDAD", "LOTE", "C.B.A", "ALIMENTO", "PRESENTACIÓN", "PRECIO UNITARIO"]
        + [d.isoformat() for d in days]
        + ["TOTAL", "Observaciones"]
    )
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = _bold()
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center")
        cell.border = _border()

    for i, row in enumerate(rows, 2):
        for c, h in enumerate(headers, 1):
            v = row.get(h)
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = _border()
            # Formato de numeros
            if h not in ("UNIDAD", "LOTE", "C.B.A", "ALIMENTO", "PRESENTACIÓN", "Observaciones"):
                cell.alignment = Alignment(horizontal="right")
                if isinstance(v, (int, float)):
                    cell.number_format = "#,##0.##"
                if h == "PRECIO UNITARIO" and isinstance(v, (int, float)):
                    cell.number_format = "$#,##0.00"

    # Anchos
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 12
    for i, _ in enumerate(days):
        ws.column_dimensions[get_column_letter(7 + i)].width = 11
    ws.column_dimensions[get_column_letter(7 + len(days))].width = 10
    ws.column_dimensions[get_column_letter(7 + len(days) + 1)].width = 18

    ws.freeze_panes = "G2"


def _write_pivot_sheet(
    ws, rows: list[dict], lote_filter: Optional[str], titulo: str,
):
    """Pivot LOTE+ALIMENTO+PRESENTACION -> suma TOTAL (filtrable por lote)."""
    ws.cell(row=1, column=1, value="UNIDAD").font = _bold()
    ws.cell(row=1, column=2, value="(Todas)")
    ws.cell(row=3, column=1, value=f"Suma de TOTAL — {titulo}").font = Font(bold=True)
    ws.cell(row=4, column=1, value="LOTE").font = _bold()
    ws.cell(row=4, column=2, value="ALIMENTO").font = _bold()
    ws.cell(row=4, column=3, value="PRESENTACIÓN").font = _bold()
    ws.cell(row=4, column=4, value="Total").font = _bold()
    for c in range(1, 5):
        ws.cell(row=4, column=c).fill = _header_fill()
        ws.cell(row=4, column=c).font = _bold()

    # Filtrar y agrupar
    by_key: dict[tuple, float] = {}
    for r in rows:
        if lote_filter and lote_filter.upper() not in (r.get("LOTE") or "").upper():
            continue
        key = (r.get("LOTE", ""), r.get("ALIMENTO", ""), r.get("PRESENTACIÓN", ""))
        try:
            t = float(r.get("TOTAL") or 0)
        except (ValueError, TypeError):
            t = 0
        by_key[key] = by_key.get(key, 0) + t

    sorted_items = sorted(by_key.items(), key=lambda x: (x[0][0], x[0][1]))
    for i, ((lote, alim, pres), tot) in enumerate(sorted_items, 5):
        ws.cell(row=i, column=1, value=lote).border = _border()
        ws.cell(row=i, column=2, value=alim).border = _border()
        ws.cell(row=i, column=3, value=pres).border = _border()
        c = ws.cell(row=i, column=4, value=tot if tot > 0 else "")
        c.number_format = "#,##0.##"
        c.border = _border()

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12


def write_base_maestra_xlsx(
    rows: list[dict],
    fecha_inicio: date,
    fecha_fin: date,
    output_path: Path,
):
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Por consistencia con la base legacy, hojas en este orden:
    # Almacen, Carne, FyV, Hoja1, BD
    ws_alm = wb.active
    ws_alm.title = "Almacen"
    ws_carne = wb.create_sheet("Carne")
    ws_fyv = wb.create_sheet("FyV")
    ws_pivot = wb.create_sheet("Hoja1")
    ws_bd = wb.create_sheet("BD")

    delta = (fecha_fin - fecha_inicio).days + 1
    days = [fecha_inicio + timedelta(days=i) for i in range(delta)]

    _write_bd_sheet(ws_bd, rows, days)

    # Pivot ALMACEN: lotes 1, 6, 7, 8, EXTRA ABARROTES (no perecederos)
    almacen_rows = [
        r for r in rows
        if any(
            kw in (r.get("LOTE") or "").upper()
            for kw in ("ABARROTE", "PAN", "BOX", "FORMULA", "FÓRMULA")
        )
    ]
    _write_pivot_sheet(ws_alm, almacen_rows, None, "Almacen")

    # Pivot CARNE: lote 2 (proteinas) + 4 (embutidos) + 3 (lacteos)
    carne_rows = [
        r for r in rows
        if any(
            kw in (r.get("LOTE") or "").upper()
            for kw in ("PROTEINA", "EMBUTIDO", "LACTEO", "LÁCTEO", "EXTRA PROTEIN")
        )
    ]
    _write_pivot_sheet(ws_carne, carne_rows, None, "Carne y lacteos")

    # Pivot FyV: lote 5
    fyv_rows = [
        r for r in rows
        if "FRUTAS Y VERDURAS" in (r.get("LOTE") or "").upper()
        or "5 FRUTAS" in (r.get("LOTE") or "").upper()
    ]
    _write_pivot_sheet(ws_fyv, fyv_rows, None, "Frutas y Verduras")

    # Pivot completo
    _write_pivot_sheet(ws_pivot, rows, None, "General")

    wb.save(str(output_path))
    return output_path


# ─── Cross-check vs anterior ───────────────────────────────────────────
def _read_bd_from_existing(path: Path) -> list[dict]:
    """Lee la hoja BD de una Base Maestra existente para comparar."""
    if not path or not path.exists():
        return []
    try:
        wb = load_workbook(path, data_only=True)
        if "BD" not in wb.sheetnames:
            return []
        ws = wb["BD"]
        headers = [str(c.value or "").strip() for c in ws[1]]
        rows = []
        for r in range(2, ws.max_row + 1):
            row = {}
            for c, h in enumerate(headers, 1):
                row[h] = ws.cell(row=r, column=c).value
            if row.get("UNIDAD") and row.get("ALIMENTO"):
                rows.append(row)
        wb.close()
        return rows
    except Exception as e:
        log.warning(f"No se pudo leer BD de {path}: {e}")
        return []


def cross_check(
    rows_new: list[dict],
    base_anterior: Optional[Path],
) -> dict:
    """Compara rows_new contra la BD anterior. Devuelve resumen de diff."""
    out = {
        "filas_nuevas": len(rows_new),
        "filas_anterior": 0,
        "diff_pct": None,
        "productos_nuevos": [],
        "productos_perdidos": [],
        "cantidad_subio_50pct": [],
        "cantidad_bajo_50pct": [],
    }
    if not base_anterior:
        return out

    rows_old = _read_bd_from_existing(base_anterior)
    out["filas_anterior"] = len(rows_old)
    if not rows_old:
        return out

    diff_count = abs(len(rows_new) - len(rows_old))
    out["diff_pct"] = round(diff_count / len(rows_old), 4) if rows_old else None

    def _key(r):
        return (
            _norm(r.get("UNIDAD") or ""),
            _norm(r.get("LOTE") or ""),
            _norm(r.get("ALIMENTO") or ""),
        )

    map_old = {_key(r): r for r in rows_old}
    map_new = {_key(r): r for r in rows_new}

    for k in list(map_new.keys())[:200]:  # cap para no explotar el output
        if k not in map_old:
            out["productos_nuevos"].append({
                "unidad": k[0], "lote": k[1], "alimento": k[2],
            })
        else:
            try:
                old_t = float(map_old[k].get("TOTAL") or 0)
                new_t = float(map_new[k].get("TOTAL") or 0)
                if old_t > 0 and new_t > old_t * 1.5:
                    out["cantidad_subio_50pct"].append({
                        "unidad": k[0], "alimento": k[2],
                        "old": old_t, "new": new_t,
                    })
                if old_t > 0 and new_t < old_t * 0.5:
                    out["cantidad_bajo_50pct"].append({
                        "unidad": k[0], "alimento": k[2],
                        "old": old_t, "new": new_t,
                    })
            except Exception:
                pass

    for k in list(map_old.keys())[:200]:
        if k not in map_new:
            out["productos_perdidos"].append({
                "unidad": k[0], "lote": k[1], "alimento": k[2],
            })

    return out


# ─── Builder principal ─────────────────────────────────────────────────
def build_base_maestra(
    source_folder: Path,
    fecha_inicio: date,
    fecha_fin: date,
    *,
    output_path: Optional[Path] = None,
    base_anterior: Optional[Path] = None,
) -> BaseMaestraResult:
    source_folder = Path(source_folder)
    if not source_folder.exists():
        raise FileNotFoundError(source_folder)

    semana_label = (
        f"{fecha_inicio.day} al {fecha_fin.day} de "
        f"{fecha_inicio.strftime('%B').lower()} {fecha_inicio.year}"
    )
    result = BaseMaestraResult(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        semana_label=semana_label,
        hospitales=[],
    )

    # Hospitales = subcarpetas que empiezan con "H. " o tienen >= 1 .xlsx
    candidate_dirs = sorted([
        d for d in source_folder.iterdir()
        if d.is_dir()
        and (d.name.startswith("H. ") or list(d.glob("*.xlsx")))
        # Filtrar carpetas que NO son hospitales (ej. "Pedido X de mayo")
        and "PEDIDO " not in _norm(d.name)
    ])

    log.info(f"Procesando {len(candidate_dirs)} carpetas de hospitales en {source_folder}")
    for d in candidate_dirs:
        try:
            r = _process_hospital_folder(d, fecha_inicio, fecha_fin)
        except Exception as e:
            r = HospitalParseResult(
                folder=d, nombre=d.name, nombre_canonico=None,
                estado="FAIL", archivos=[], items=[],
                errors=[f"Excepcion: {e}"],
            )
        result.hospitales.append(r)

    # Generar BD rows
    rows = _make_bd_rows(result.hospitales, fecha_inicio, fecha_fin)
    result.filas_bd = len(rows)

    # Cross-check si hay anterior
    if base_anterior and base_anterior.exists():
        try:
            cross = cross_check(rows, base_anterior)
            result.diff_resumen = cross
            result.diff_pct_vs_anterior = cross.get("diff_pct")
        except Exception as e:
            log.warning(f"Cross-check fallo: {e}")

    # Generar Excel
    output_path = output_path or Path(
        f"/tmp/cadena_base_maestra/Base maestra {fecha_inicio.isoformat()} al {fecha_fin.isoformat()}.xlsx"
    )
    write_base_maestra_xlsx(rows, fecha_inicio, fecha_fin, output_path)
    result.output_path = output_path
    result.output_size = output_path.stat().st_size if output_path.exists() else 0

    return result
