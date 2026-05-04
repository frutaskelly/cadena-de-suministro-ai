"""Parser autodetect de archivos de pedido hospitalario (Sprint 13).

Reconoce el formato APENDICE 4 - ORDEN DE SUMINISTRO con tolerancia a:
- Hojas multiples (SEMANA 1..5) -> escoge la que corresponde a la fecha
- Header en posicion variable (fila 15-25)
- Hospital extraido del cuerpo del archivo
- Cantidades vacias / strings / numeros
- Acentos / encoding mixto

Output normalizado por hospital:
{
  "archivo": "/path/...",
  "estado": "OK" | "WARN" | "FAIL",
  "hospital_nombre_archivo": "...",
  "hospital_canonico": "..." (si matchea catalogo),
  "fecha_inicio": date | None,
  "fecha_fin": date | None,
  "items": [{lote, cba, alimento, presentacion, precio, dias: {2026-05-04: 4, ...}}],
  "warnings": [...],
  "errors": [...],
}
"""
from __future__ import annotations

import logging
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# Header keywords — tolerantes a variaciones
HEADER_KEYS_LOTE = {"LOTE", "LOTE "}
HEADER_KEYS_CBA = {"C.B.A", "CBA", "C.B.A."}
HEADER_KEYS_ALIMENTO = {
    "ALIMENTO", "NOMBRE DEL ALIMENTO", "NOMBRE DEL ALIMENT",
    "PRODUCTO", "DESCRIPCION", "DESCRIPCIÓN",
}
HEADER_KEYS_PRES = {"PRESENTACION", "PRESENTACIÓN", "PRESENTACIO"}
HEADER_KEYS_PRECIO = {
    "PRECIO UNITARIO", "PRECIO UNITARIO ", "PRECIO",
    "P. UNITARIO", "P/U",
}
HEADER_KEYS_TOTAL = {"TOTAL", "TOTAL "}


def _norm(s) -> str:
    """Normaliza string: upper + strip accents para comparacion."""
    if s is None:
        return ""
    s = str(s).strip().upper()
    # Quitar acentos
    s = "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )
    return s


def _is_header_row(row_values: list) -> bool:
    """True si la fila parece ser el header de la tabla de productos."""
    norm = [_norm(v) for v in row_values]
    has_lote = any(_norm(v) in {_norm(k) for k in HEADER_KEYS_LOTE} for v in row_values)
    has_alimento = any(_norm(v) in {_norm(k) for k in HEADER_KEYS_ALIMENTO} for v in row_values)
    has_presentacion = any(_norm(v) in {_norm(k) for k in HEADER_KEYS_PRES} for v in row_values)
    return has_lote and has_alimento and has_presentacion


def _find_header_row(ws, max_scan: int = 30) -> Optional[int]:
    """Devuelve el nro de fila del header (1-indexed) o None."""
    last_col = min(ws.max_column, 25)
    for r in range(1, min(max_scan, ws.max_row) + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, last_col + 1)]
        if _is_header_row(row):
            return r
    return None


def _map_header_columns(ws, header_row: int) -> dict[str, int]:
    """Mapea cada columna del header a un campo conocido. 1-indexed."""
    mapping: dict[str, int] = {}
    last_col = min(ws.max_column, 25)
    for c in range(1, last_col + 1):
        v = _norm(ws.cell(row=header_row, column=c).value)
        if not v:
            continue
        if v in {_norm(k) for k in HEADER_KEYS_LOTE}:
            mapping["lote"] = c
        elif v in {_norm(k) for k in HEADER_KEYS_CBA}:
            mapping["cba"] = c
        elif v in {_norm(k) for k in HEADER_KEYS_ALIMENTO}:
            mapping["alimento"] = c
        elif v in {_norm(k) for k in HEADER_KEYS_PRES}:
            mapping["presentacion"] = c
        elif v in {_norm(k) for k in HEADER_KEYS_PRECIO}:
            mapping["precio"] = c
        elif v in {_norm(k) for k in HEADER_KEYS_TOTAL}:
            mapping["total"] = c
    return mapping


def _detect_day_columns(ws, header_row: int) -> dict[date, int]:
    """Detecta columnas de dias.

    Estrategia: en el header_row, las columnas con valor numerico 1-31 son dias.
    Si no hay numeros, busca en filas anteriores (header_row-1, -2) por
    encabezados tipo 'L | MA | MI | J | V | S | D' o fechas.
    Devuelve dict: {date(2026,5,4): col_idx, ...}.
    """
    last_col = min(ws.max_column, 25)
    days: dict[date, int] = {}

    # 1) Header fila contiene numeros 1-31? (admite int, float, string,
    # o expresion tipo "DIA 9", "DÍA 4", "D-5", "Dia 12")
    header_nums: dict[int, int] = {}
    for c in range(1, last_col + 1):
        v = ws.cell(row=header_row, column=c).value
        n = None
        if isinstance(v, (int, float)) and 1 <= v <= 31:
            n = int(v)
        elif isinstance(v, str):
            s = v.strip()
            if s.isdigit():
                vi = int(s)
                if 1 <= vi <= 31:
                    n = vi
            else:
                # Extraer numero de "DÍA 9" / "D-5" / "Dia 12"
                m = re.search(r"\b(\d{1,2})\b", s)
                if m and ("DIA" in _norm(s) or "DÍA" in s.upper()):
                    vi = int(m.group(1))
                    if 1 <= vi <= 31:
                        n = vi
        if n is not None:
            header_nums[c] = n

    # Si tenemos suficientes numeros (>= 5 dias seguidos), buscar la fecha
    # base de fila anterior (mes / año)
    fecha_base: Optional[date] = None
    if len(header_nums) >= 5:
        # buscar fila con mes/año en filas previas
        for r in range(max(1, header_row - 5), header_row):
            for c in range(1, last_col + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, datetime):
                    fecha_base = date(v.year, v.month, 1)
                    break
                if isinstance(v, str):
                    m = re.search(r"(20\d{2})", v)
                    if m:
                        # buscar mes en la misma celda
                        for mes_num, mes_kw in enumerate([
                            "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO",
                            "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE",
                            "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
                        ], 1):
                            if mes_kw in _norm(v):
                                fecha_base = date(int(m.group(1)), mes_num, 1)
                                break
                    if fecha_base:
                        break
            if fecha_base:
                break

        # Si no detecto fecha_base, asume "current month" como fallback
        if not fecha_base:
            today = date.today()
            fecha_base = date(today.year, today.month, 1)

        for col, day_num in header_nums.items():
            try:
                d = date(fecha_base.year, fecha_base.month, day_num)
                days[d] = col
            except ValueError:
                continue
        return days

    # 2) Header tiene fechas como datetime
    for c in range(1, last_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, datetime):
            days[v.date()] = c

    return days


def _extract_hospital_name(ws, header_row: int) -> Optional[str]:
    """Busca 'Hospital y/o Unidad' en filas previas al header y extrae el valor."""
    last_col = min(ws.max_column, 15)
    for r in range(1, header_row):
        for c in range(1, last_col + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            n = _norm(v)
            if "HOSPITAL Y/O UNIDAD" in n or "HOSPITAL Y O UNIDAD" in n or "UNIDAD HOSPITAL" in n:
                # el valor del hospital esta tipicamente 1-3 columnas a la derecha
                for cc in range(c + 1, min(c + 6, last_col + 1)):
                    val = ws.cell(row=r, column=cc).value
                    if isinstance(val, str) and len(val.strip()) > 5:
                        return val.strip()
    # Fallback: titulo grande en filas 1-7
    for r in range(1, min(8, header_row)):
        for c in range(1, last_col + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and "HOSPITAL" in _norm(v) and len(v) > 15:
                return v.strip()
    return None


def _to_decimal(v) -> Optional[Decimal]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float, Decimal)):
        try:
            return Decimal(str(v))
        except (InvalidOperation, ValueError):
            return None
    if isinstance(v, str):
        s = v.strip().replace(",", ".")
        if not s:
            return None
        try:
            return Decimal(s)
        except (InvalidOperation, ValueError):
            return None
    return None


def _pick_sheet(
    wb,
    fecha_inicio: Optional[date],
    fecha_fin: Optional[date],
) -> str:
    """Elige la hoja a parsear cuando hay multiples.

    Estrategia (en orden):
    1. Si solo hay 1 hoja con header detectable, esa.
    2. Si fecha_inicio + fecha_fin se pasan, escoge la hoja cuyos dias
       cubran MEJOR el rango (max overlap con [fecha_inicio, fecha_fin]).
    3. Si solo fecha_inicio, escoge hoja que incluya esa fecha.
    4. Default: la hoja con MAS items (cantidades > 0) entre las que
       tienen header detectable. Esto evita escoger hojas auxiliares
       como DESECHABLES si la principal tiene más data.
    """
    sheets = wb.sheetnames

    candidates_with_header = []
    for sn in sheets:
        ws = wb[sn]
        hr = _find_header_row(ws)
        if hr:
            candidates_with_header.append((sn, hr))

    if not candidates_with_header:
        return sheets[0]
    if len(candidates_with_header) == 1:
        return candidates_with_header[0][0]

    # Si tenemos rango de fechas, escoger por overlap
    if fecha_inicio:
        best_sn = None
        best_score = -1
        for sn, hr in candidates_with_header:
            ws = wb[sn]
            days = _detect_day_columns(ws, hr)
            if not days:
                continue
            score = 0
            for d in days:
                if fecha_fin:
                    if fecha_inicio <= d <= fecha_fin:
                        score += 1
                elif d == fecha_inicio:
                    score += 10  # match exacto al inicio
                elif min(days) <= fecha_inicio <= max(days):
                    score += 1
            if score > best_score:
                best_score = score
                best_sn = sn
        if best_sn and best_score > 0:
            return best_sn

    # Default: la que tenga mas items con cantidades > 0 en columnas de dia
    # o columna TOTAL. Si no hay days/total mapeados, usa fallback de
    # numeros >0 en cualquier columna a la derecha del precio.
    best_sn = candidates_with_header[0][0]
    best_count = -1
    for sn, hr in candidates_with_header:
        ws = wb[sn]
        cols = _map_header_columns(ws, hr)
        if "lote" not in cols or "alimento" not in cols:
            continue
        days = _detect_day_columns(ws, hr)
        # columnas a contar: dias detectados + TOTAL si esta
        cols_a_contar = list((days or {}).values())
        if "total" in cols:
            cols_a_contar.append(cols["total"])
        # Si no hay nada, fallback: cualquier columna a la derecha del header
        # de "PRECIO" (donde tipicamente vendrian las cantidades)
        if not cols_a_contar:
            start = max(cols.get("precio", 0), cols.get("presentacion", 0)) + 1
            cols_a_contar = list(range(start, min(ws.max_column, 20) + 1))

        count = 0
        for r in range(hr + 1, min(hr + 400, ws.max_row + 1)):
            if not ws.cell(row=r, column=cols.get("alimento", 1)).value:
                continue
            for c in cols_a_contar:
                v = ws.cell(row=r, column=c).value
                if isinstance(v, (int, float)) and v > 0:
                    count += 1
                    break
                if isinstance(v, str):
                    s = v.strip().replace(",", ".")
                    try:
                        if float(s) > 0:
                            count += 1
                            break
                    except (ValueError, TypeError):
                        pass
        if count > best_count:
            best_count = count
            best_sn = sn
    return best_sn


def parse_pedido_hospital(
    xlsx_path: Path,
    *,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    # alias legacy
    fecha_objetivo: Optional[date] = None,
) -> dict:
    """Parsea un archivo de pedido hospitalario.

    Args:
        xlsx_path: ruta al .xlsx
        fecha_inicio, fecha_fin: rango objetivo. Filtra dias del output
            a [fecha_inicio, fecha_fin] y escoge la hoja correcta entre
            multiples SEMANA N.
        fecha_objetivo: alias legacy de fecha_inicio.

    Returns dict con shape:
        {
          "archivo": "/path/...",
          "estado": "OK" | "WARN" | "FAIL",
          "hospital_nombre_archivo": "...",
          "hojas_disponibles": [...],
          "hoja_usada": "...",
          "fecha_inicio": date | None,
          "fecha_fin": date | None,
          "items": [...],
          "warnings": [...],
          "errors": [...],
        }
    """
    xlsx_path = Path(xlsx_path)
    result = {
        "archivo": str(xlsx_path),
        "estado": "OK",
        "hospital_nombre_archivo": xlsx_path.parent.name,
        "hojas_disponibles": [],
        "hoja_usada": None,
        "fecha_inicio": None,
        "fecha_fin": None,
        "items": [],
        "warnings": [],
        "errors": [],
    }

    if not xlsx_path.exists():
        result["estado"] = "FAIL"
        result["errors"].append(f"Archivo no existe: {xlsx_path}")
        return result

    if xlsx_path.suffix.lower() not in (".xlsx", ".xlsm"):
        result["estado"] = "FAIL"
        result["errors"].append(f"Tipo no soportado: {xlsx_path.suffix} (PDFs requieren OCR)")
        return result

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        result["estado"] = "FAIL"
        result["errors"].append(f"Error abriendo Excel: {e}")
        return result

    fecha_inicio = fecha_inicio or fecha_objetivo
    result["hojas_disponibles"] = wb.sheetnames
    sheet_name = _pick_sheet(wb, fecha_inicio, fecha_fin)
    result["hoja_usada"] = sheet_name
    ws = wb[sheet_name]

    header_row = _find_header_row(ws)
    if not header_row:
        result["estado"] = "FAIL"
        result["errors"].append(
            f"No se detecto header en hoja '{sheet_name}' (rev. filas 1-30)"
        )
        return result

    cols = _map_header_columns(ws, header_row)
    if "lote" not in cols or "alimento" not in cols:
        result["estado"] = "FAIL"
        result["errors"].append(
            f"Header en fila {header_row} sin columnas requeridas. Mapeo: {cols}"
        )
        return result

    days = _detect_day_columns(ws, header_row)
    if not days:
        result["warnings"].append(
            f"No se detectaron columnas de dia en fila {header_row}; "
            f"se generara TOTAL pero sin desglose por dia"
        )

    if days:
        result["fecha_inicio"] = min(days).isoformat()
        result["fecha_fin"] = max(days).isoformat()

    # Hospital
    hospital_extraido = _extract_hospital_name(ws, header_row)
    if hospital_extraido:
        result["hospital_nombre_archivo"] = hospital_extraido

    # Iterate filas de datos
    items: list[dict] = []
    for r in range(header_row + 1, ws.max_row + 1):
        lote = ws.cell(row=r, column=cols["lote"]).value
        alimento = ws.cell(row=r, column=cols["alimento"]).value

        if not lote and not alimento:
            continue
        if not alimento or not str(alimento).strip():
            continue

        cba = (
            str(ws.cell(row=r, column=cols["cba"]).value or "").strip()
            if "cba" in cols else ""
        )
        presentacion = (
            str(ws.cell(row=r, column=cols["presentacion"]).value or "").strip()
            if "presentacion" in cols else ""
        )
        precio = (
            _to_decimal(ws.cell(row=r, column=cols["precio"]).value)
            if "precio" in cols else None
        )

        dia_qty: dict[str, float] = {}
        total_row = Decimal(0)
        for d, c in days.items():
            # Filtrar al rango [fecha_inicio, fecha_fin]
            if fecha_inicio and d < fecha_inicio:
                continue
            if fecha_fin and d > fecha_fin:
                continue
            qty = _to_decimal(ws.cell(row=r, column=c).value)
            if qty is not None and qty > 0:
                dia_qty[d.isoformat()] = float(qty)
                total_row += qty

        # Si no tiene dias pero tiene total > 0, lo ponemos en lote_total
        total_col_val = (
            _to_decimal(ws.cell(row=r, column=cols["total"]).value)
            if "total" in cols else None
        )

        if not dia_qty and not total_col_val:
            # fila vacia (catalogo sin pedido)
            continue

        items.append({
            "lote": str(lote).strip() if lote else "",
            "cba": cba,
            "alimento": str(alimento).strip(),
            "presentacion": presentacion,
            "precio": float(precio) if precio is not None else None,
            "dias": dia_qty,
            "total": float(total_row) if total_row > 0 else (
                float(total_col_val) if total_col_val else 0.0
            ),
        })

    result["items"] = items

    if not items:
        result["estado"] = "WARN"
        result["warnings"].append("No se encontraron lineas con cantidades > 0")

    return result
