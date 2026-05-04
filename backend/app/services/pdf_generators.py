"""Generadores de PDF para pedidos del dia (port v1).

- generar_pedido_pdf: una pagina por hospital, sin precios.
- generar_lista_compras_pdf: consolidado de compras al mayoreo.
- generar_lista_compras_xlsx: misma data en Excel.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

log = logging.getLogger(__name__)

AZUL_OSC = colors.HexColor("#1F4E79")
AZUL_CLAR = colors.HexColor("#D6E4F0")
GRIS = colors.HexColor("#BBBBBB")
GRIS_CLARO = colors.HexColor("#F8F9FB")


def _fmt_qty(v) -> str:
    try:
        f = float(v)
    except (TypeError, ValueError):
        return str(v)
    return str(int(f)) if f == int(f) else f"{f:.2f}"


# ─── Pedido por hospital ───────────────────────────────────────────────────
def generar_pedido_pdf(
    items: Iterable[dict],
    fecha_str: str,
    output_path: Path,
    *,
    subtitulo: str = "Lote 5: Frutas y Verduras",
    titulo_principal: Optional[str] = None,
) -> Path:
    """Una pagina por unidad de entrega. Sin precios.

    Items: lista de dicts {unidad, alimento, presentacion, cantidad}.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Agrupar por unidad
    por_unidad: dict[str, list[dict]] = {}
    for it in items:
        unidad = it["unidad"]
        por_unidad.setdefault(unidad, []).append(it)

    # Sumar cantidad por (alimento, presentacion) dentro de cada unidad
    for unidad, lista in por_unidad.items():
        agg: dict[tuple[str, str], float] = {}
        for it in lista:
            key = (it["alimento"], it["presentacion"])
            agg[key] = agg.get(key, 0) + float(it.get("cantidad", 0))
        por_unidad[unidad] = [
            {"alimento": k[0], "presentacion": k[1], "cantidad": v}
            for k, v in agg.items()
            if v > 0
        ]
        por_unidad[unidad].sort(key=lambda r: r["alimento"])

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=letter,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        title=f"Pedido {fecha_str}",
    )

    styles = getSampleStyleSheet()
    style_h1 = ParagraphStyle(
        "h1", parent=styles["Heading1"],
        fontSize=14, leading=18, textColor=AZUL_OSC, spaceAfter=4,
    )
    style_sub = ParagraphStyle(
        "sub", parent=styles["Normal"],
        fontSize=10, leading=12,
        textColor=colors.HexColor("#555555"),
        spaceAfter=12,
    )
    style_banner = ParagraphStyle(
        "banner", parent=styles["Heading2"],
        fontSize=11, leading=13,
        textColor=colors.HexColor("#7C3AED"),
        fontName="Helvetica-Bold", spaceAfter=4,
    )

    elements = []
    unidades = sorted(por_unidad.keys())

    for i, unidad in enumerate(unidades):
        rows = por_unidad[unidad]
        if not rows:
            continue

        if titulo_principal:
            elements.append(Paragraph(titulo_principal, style_banner))
        elements.append(Paragraph(unidad, style_h1))
        elements.append(
            Paragraph(
                f"Pedido del <b>{fecha_str}</b> &nbsp;·&nbsp; {subtitulo}",
                style_sub,
            )
        )

        data = [["#", "Alimento", "Presentación", "Cantidad"]]
        for idx, r in enumerate(rows, 1):
            data.append([
                str(idx), r["alimento"], r["presentacion"], _fmt_qty(r["cantidad"]),
            ])
        total = sum(r["cantidad"] for r in rows)
        data.append(["", "TOTAL DE PRODUCTOS", str(len(rows)), _fmt_qty(total)])

        table = Table(
            data,
            colWidths=[0.9 * cm, 10.5 * cm, 4 * cm, 2.5 * cm],
            repeatRows=1,
        )
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), AZUL_OSC),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (3, 0), (3, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("FONTNAME", (0, 1), (-1, -2), "Helvetica"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.4, GRIS),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, GRIS_CLARO]),
            ("BACKGROUND", (0, -1), (-1, -1), AZUL_CLAR),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ]))
        elements.append(table)
        if i < len(unidades) - 1:
            elements.append(PageBreak())

    doc.build(elements)
    log.info(f"Pedido PDF generado: {output_path.name} ({len(unidades)} unidades)")
    return output_path


# ─── Lista de compras consolidada ─────────────────────────────────────────
def _consolidar_items(items: Iterable[dict]) -> list[dict]:
    agg: dict[tuple[str, str], float] = {}
    for it in items:
        key = (it["alimento"], it["presentacion"])
        agg[key] = agg.get(key, 0) + float(it.get("cantidad", 0))
    return [
        {"alimento": k[0], "presentacion": k[1], "cantidad": v}
        for k, v in sorted(agg.items())
        if v > 0
    ]


def generar_lista_compras_pdf(
    items: Iterable[dict],
    fecha_str: str,
    output_path: Path,
) -> Path:
    """PDF consolidado de compras al mayoreo (sin desglose por hospital)."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    consolidado = _consolidar_items(items)

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=letter,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        title=f"Lista de Compras {fecha_str}",
    )
    styles = getSampleStyleSheet()
    style_h1 = ParagraphStyle(
        "h1", parent=styles["Heading1"],
        fontSize=16, leading=20, textColor=AZUL_OSC,
        alignment=TA_CENTER, spaceAfter=4,
    )
    style_sub = ParagraphStyle(
        "sub", parent=styles["Normal"],
        fontSize=10, leading=12,
        textColor=colors.HexColor("#555555"),
        alignment=TA_CENTER, spaceAfter=18,
    )

    elements = [
        Paragraph("Lista de Compras al Mayoreo", style_h1),
        Paragraph(
            f"Pedido del <b>{fecha_str}</b> &nbsp;·&nbsp; "
            f"{len(consolidado)} productos",
            style_sub,
        ),
    ]

    data = [["#", "Alimento", "Presentación", "Cantidad"]]
    for idx, r in enumerate(consolidado, 1):
        data.append([
            str(idx), r["alimento"], r["presentacion"], _fmt_qty(r["cantidad"]),
        ])
    total = sum(r["cantidad"] for r in consolidado)
    data.append(["", "TOTAL", str(len(consolidado)), _fmt_qty(total)])

    table = Table(
        data,
        colWidths=[1 * cm, 10.5 * cm, 4 * cm, 2.5 * cm],
        repeatRows=1,
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), AZUL_OSC),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (3, 0), (3, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("FONTNAME", (0, 1), (-1, -2), "Helvetica"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.4, GRIS),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, GRIS_CLARO]),
        ("BACKGROUND", (0, -1), (-1, -1), AZUL_CLAR),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    elements.append(table)
    doc.build(elements)
    log.info(f"Lista compras PDF: {output_path.name}")
    return output_path


def generar_lista_compras_xlsx(
    items: Iterable[dict],
    fecha_str: str,
    output_path: Path,
) -> Path:
    """Excel con la lista consolidada para compras."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    consolidado = _consolidar_items(items)

    wb = Workbook()
    ws = wb.active
    ws.title = "Lista de Compras"

    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    border = Border(
        left=Side(border_style="thin", color="BBBBBB"),
        right=Side(border_style="thin", color="BBBBBB"),
        top=Side(border_style="thin", color="BBBBBB"),
        bottom=Side(border_style="thin", color="BBBBBB"),
    )

    # Title row
    ws["A1"] = f"Lista de Compras al Mayoreo — {fecha_str}"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 28

    # Header
    headers = ["#", "Alimento", "Presentación", "Cantidad"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = header_font
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Body
    for i, r in enumerate(consolidado, 1):
        ws.cell(row=3 + i, column=1, value=i).border = border
        ws.cell(row=3 + i, column=2, value=r["alimento"]).border = border
        ws.cell(row=3 + i, column=3, value=r["presentacion"]).border = border
        c = ws.cell(row=3 + i, column=4, value=float(r["cantidad"]))
        c.border = border
        c.alignment = Alignment(horizontal="right")

    # Total row
    last = 3 + len(consolidado) + 1
    ws.cell(row=last, column=2, value="TOTAL").font = Font(bold=True)
    ws.cell(row=last, column=3, value=len(consolidado)).font = Font(bold=True)
    ws.cell(row=last, column=4, value=sum(r["cantidad"] for r in consolidado)).font = Font(bold=True)
    for col in range(1, 5):
        ws.cell(row=last, column=col).fill = PatternFill(
            start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"
        )
        ws.cell(row=last, column=col).border = border

    # Column widths
    widths = {1: 5, 2: 40, 3: 18, 4: 12}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    wb.save(str(output_path))
    log.info(f"Lista compras XLSX: {output_path.name}")
    return output_path
